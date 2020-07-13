import requests
import json
from openpyxl import Workbook


wb = Workbook()
ws=wb.active
ws1 = wb.create_sheet("sheet2")

api_key="746cda8bd2e921231b7461c2ace22b88"
r=requests.get("http://api.openweathermap.org/data/2.5/group?id=1264733,1264728&appid=746cda8bd2e921231b7461c2ace22b88").json()
for i in r['list']:
    print(i['name'])

ws['A1']='City Name'
ws['B1']='Temperature'
ws['C1']='Temperature in C'
ws['D1']='Temperature in F'
x=1
for i in r['list']:
    ws.cell(row=x+1,column=1,value=i['name'])
    x=x+1
y=1
for i in r['list']:
    ws.cell(row=y+1,column=2,value=i['main']['temp'])
    ws.cell(row=y+1,column=3,value=(i['main']['temp']-273.15))
    ws.cell(row=y+1,column=4,value=(i['main']['temp']-273.15)*1.8+32)
    
    y=y+1

    
ws1['A1']='City name'
with open('world-cities_json.json')as data:
    result=json.load(data)
z=1
for i in result:
    ws1.cell(row=z+1,column=1,value=i['name'])
    z=z+1

    
wb.save('document.xlsx')


        
